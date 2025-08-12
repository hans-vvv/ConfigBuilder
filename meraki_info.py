from collections import defaultdict
from functools import cached_property

import openpyxl

import meraki
from meraki.exceptions import APIError

from credentials import API_KEY
from utils import xlref, custom_layout_sheet


class MerakiInfo:

    """
    This class exposes methods to get network information out of the Meraki SD-WAN solution using REST API
    and HTTP scraping if REST API is not sufficient and prints the data in an Excel file.
    """

    WORKBOOK_NAME = 'result.xlsx'

    def __init__(self):
        self.dashboard = None
        self.wb = openpyxl.Workbook()

        self._create_meraki_dashboard_api()

    def _create_meraki_dashboard_api(self):
        self.dashboard = meraki.DashboardAPI(api_key=API_KEY, print_console=False, suppress_logging=True)

    @cached_property
    def _get_network_id_by_name_cache(self):

        cache = {}
        orgs = self.dashboard.organizations.getOrganizations()
        for org in orgs:
            networks = self.dashboard.organizations.getOrganizationNetworks(org['id'])
            for network in networks:
                cache[network['name']] = network['id']
        return cache

    def get_network_id_by_name(self, network_name):
        return self._get_network_id_by_name_cache[network_name]

    @cached_property
    def _get_network_appliance_static_routes_cache(self):

        cache = {}
        network_names = self._get_network_names
        for network_name in network_names:
            network_id = self.get_network_id_by_name(network_name)
            try:
                static_routes = self.dashboard.appliance.getNetworkApplianceStaticRoutes(network_id)
                cache[network_name] = static_routes
            except APIError:
                pass
        return cache

    def _get_fixed_ip_assignments(self):
        result = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
        network_names = self._get_network_names
        for network_name in network_names:
            subnets = self._get_network_appliance_static_routes_cache.get(network_name)
            if subnets:
                for subnet in subnets:
                    # print(subnet)
                    subnet_name = subnet.get('name')
                    enabled = subnet.get('enabled')
                    fixed_assignments = subnet.get('fixedIpAssignments')
                    if fixed_assignments:
                        for mac, vals in fixed_assignments.items():
                            fixed_ip = vals['ip']
                            name = vals['name']
                            result[network_name][subnet_name][fixed_ip]['mac'] = mac
                            result[network_name][subnet_name][fixed_ip]['name'] = name
                            result[network_name][subnet_name][fixed_ip]['enabled'] = enabled

        return result

    def print_fixed_ip_assignments_to_excel_tab(self, sheet_name):

        fixed_ip_assignments = self._get_fixed_ip_assignments()

        self.wb.create_sheet(sheet_name)
        sheet = self.wb[sheet_name]

        index = 0
        sheet[xlref(0, 0)] = 'Network Name'
        sheet[xlref(0, 1)] = 'Subnet Name'
        sheet[xlref(0, 2)] = 'IP address'
        sheet[xlref(0, 3)] = 'MAC address'
        sheet[xlref(0, 4)] = 'Description'
        sheet[xlref(0, 5)] = 'Enabled'

        for network_name in fixed_ip_assignments:
            for subnet_name in fixed_ip_assignments[network_name]:
                for ip in fixed_ip_assignments[network_name][subnet_name]:

                    mac = fixed_ip_assignments[network_name][subnet_name][ip]['mac']
                    name = fixed_ip_assignments[network_name][subnet_name][ip]['name']
                    enabled = fixed_ip_assignments[network_name][subnet_name][ip]['enabled']

                    sheet[xlref(index + 1, 0)] = network_name
                    sheet[xlref(index + 1, 1)] = subnet_name
                    sheet[xlref(index + 1, 2)] = ip
                    sheet[xlref(index + 1, 3)] = mac
                    sheet[xlref(index + 1, 4)] = name
                    sheet[xlref(index + 1, 5)] = enabled
                    index += 1
        custom_layout_sheet(sheet)
        self._excel_save()

    def _get_reserved_ip_ranges(self):

        result = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
        network_names = self._get_network_names
        for network_name in network_names:
            subnets = self._get_network_appliance_static_routes_cache.get(network_name)
            if subnets:
                for subnet in subnets:
                    # print(subnet)
                    enabled = subnet.get('enabled')
                    subnet_name = subnet.get('name')
                    reserved_ip_ranges = subnet.get('reservedIpRanges')
                    if reserved_ip_ranges:
                        for reserved_ip_range in reserved_ip_ranges:
                            result[network_name][subnet_name]['res'].append(reserved_ip_range)
                    result[network_name][subnet_name]['enabled'].append(enabled)
        # print(result)
        return result

    def print_reserved_ip_ranges_to_excel_tab(self, sheet_name):

        reserved_ip_ranges = self._get_reserved_ip_ranges()

        self.wb.create_sheet(sheet_name)
        sheet = self.wb[sheet_name]

        index = 0
        sheet[xlref(0, 0)] = 'Network Name'
        sheet[xlref(0, 1)] = 'Subnet Name'
        sheet[xlref(0, 2)] = 'Start'
        sheet[xlref(0, 3)] = 'End'
        sheet[xlref(0, 4)] = 'Start'
        sheet[xlref(0, 5)] = 'End'
        sheet[xlref(0, 6)] = 'Enabled'

        for network_name in reserved_ip_ranges:
            for subnet_name in reserved_ip_ranges[network_name]:
                sheet[xlref(index + 1, 0)] = network_name
                sheet[xlref(index + 1, 1)] = subnet_name
                if len(reserved_ip_ranges[network_name][subnet_name]['res']) == 1:
                    sheet[xlref(index + 1, 2)] = reserved_ip_ranges[network_name][subnet_name]['res'][0]['start']
                    sheet[xlref(index + 1, 3)] = reserved_ip_ranges[network_name][subnet_name]['res'][0]['end']
                else:
                    for i, reserved_ip_range in enumerate(reserved_ip_ranges[network_name][subnet_name]['res']):
                        if i == 0:
                            sheet[xlref(index + 1, 2)] = reserved_ip_range['start']
                            sheet[xlref(index + 1, 3)] = reserved_ip_range['end']
                        else:
                            sheet[xlref(index + 1, 4)] = reserved_ip_range['start']
                            sheet[xlref(index + 1, 5)] = reserved_ip_range['end']
                sheet[xlref(index + 1, 6)] = reserved_ip_ranges[network_name][subnet_name]['enabled'][0]
                index += 1
        custom_layout_sheet(sheet)
        self._excel_save()

    def _get_static_routes(self):

        result = defaultdict(lambda: defaultdict(dict))
        network_names = self._get_network_names
        for network_name in network_names:
            static_routes = self._get_network_appliance_static_routes_cache.get(network_name)
            if static_routes:
                for route in static_routes:
                    result[network_name][route['name']]['subnet'] = route['subnet']
                    result[network_name][route['name']]['enabled'] = route['enabled']
        return result

    def print_static_routes_to_excel_tab(self, sheet_name):

        static_routes = self._get_static_routes()

        self.wb.create_sheet(sheet_name)
        sheet = self.wb[sheet_name]

        index = 0
        sheet[xlref(0, 0)] = 'Network name'
        sheet[xlref(0, 1)] = 'Subnet name'
        sheet[xlref(0, 2)] = 'Subnet'
        sheet[xlref(0, 3)] = 'Enabled'

        for network_name in static_routes:
            for subnet_name in static_routes[network_name]:
                sheet[xlref(index + 1, 0)] = network_name
                sheet[xlref(index + 1, 1)] = subnet_name
                sheet[xlref(index + 1, 2)] = static_routes[network_name][subnet_name]['subnet']
                sheet[xlref(index + 1, 3)] = static_routes[network_name][subnet_name]['enabled']
                index += 1
        custom_layout_sheet(sheet)
        self._excel_save()

    @cached_property
    def _get_network_names(self):
        network_names = []
        orgs = self.dashboard.organizations.getOrganizations()
        for org in orgs:
            networks = self.dashboard.organizations.getOrganizationNetworks(org['id'])
            for network in networks:
                network_names.append(network['name'])
        return network_names

    def _excel_save(self):
        self.wb.save(self.WORKBOOK_NAME)
