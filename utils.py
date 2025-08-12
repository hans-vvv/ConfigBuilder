from openpyxl import worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def xlref(row: int, column: int, zero_indexed: bool = True) -> str:
    """
    openpyxl helper to generate Excel cell references.

    Args:
        row (int): Row index (zero-indexed or one-indexed based on zero_indexed).
        column (int): Column index (zero-indexed or one-indexed based on zero_indexed).
        zero_indexed (bool, optional): Whether the row and column indices are zero-indexed.
                                      Defaults to True.

    Returns:
        str: Excel cell reference (e.g., "A1", "B2").
    """
    if zero_indexed:
        row += 1
        column += 1
    return get_column_letter(column) + str(row)


def custom_layout_sheet(sheet: worksheet) -> None:
    """
    Openpyxl helper to apply a custom layout to a worksheet.

    This function:
        - Freezes the first row.
        - Adds a filter to the entire sheet.
        - Auto-sizes columns based on cell content.
        - Makes the first row bold.
    """
    for i in range(0, sheet.max_column + 1):
        sheet.freeze_panes = xlref(1, i)

    sheet.auto_filter.ref = sheet.dimensions

    for letter in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(letter)
        max_width = 0
        for cell in sheet[column_letter]:
            if cell.value:
                max_width = max(max_width, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = (max_width + 1) * 1.50

    for cell in sheet[1]:  # Make first row bold
        cell.font = Font(bold=True)